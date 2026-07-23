#!/usr/bin/env python3
"""
YouTube 提文案 — 输入 YouTube 链接提取视频文字稿
=========================================
接口：POST /story/api/youtube/transcript

Usage:
    python3 extract.py "https://www.youtube.com/watch?v=dQw4w9WgXcQ"
    python3 extract.py "dQw4w9WgXcQ" --language "zh,en"
    python3 extract.py "https://youtu.be/dQw4w9WgXcQ" --timestamp   # 带时间戳
    python3 extract.py "URL" --excel                               # 同时导出 Excel
    python3 extract.py "URL" --json --no-metadata
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    from deep_translator import GoogleTranslator
    HAS_TRANSLATOR = True
except ImportError:
    HAS_TRANSLATOR = False

# ─── 配置 ─────────────────────────────────────────────────────────────────────────
API_URL = "https://redfox.hk/story/api/youtube/transcript"
CONFIG_FILE = Path.home() / ".qoder" / "apis" / "redfox.json"
ENV_KEY = "REDFOX_API_KEY"
SOURCE = "YouTube提文案-GitHub"

DEFAULT_OUTPUT_DIR = Path.home() / "Downloads" / "QoderYoutubeDigest"
DEFAULT_LANGUAGE = "zh,en,asr"
MAX_RETRIES = 3

# 中文语言标识前缀（API 返回的 language 字段可能为 zh / zh-CN / zh-TW 等）
ZH_PREFIX = ("zh",)

# ─── 终端颜色 ──────────────────────────────────────────────────────────────────────
GREEN = "\033[92m"
YELLOW = "\033[93m"
RED = "\033[91m"
CYAN = "\033[96m"
BOLD = "\033[1m"
DIM = "\033[2m"
RESET = "\033[0m"


def is_chinese(language):
    """判断 API 返回的 language 是否为中文"""
    lang_lower = (language or "").lower().replace("_", "-")
    return any(lang_lower.startswith(p) for p in ZH_PREFIX)


def translate_lines(lines, batch_size=30):
    """
    将非中文文案行批量翻译为中文。
    每 batch_size 行合并为一次翻译请求，减少 API 调用次数。
    返回翻译后的行列表（与输入一一对应）。
    """
    if not HAS_TRANSLATOR:
        warn("未安装 deep-translator，无法自动翻译，请执行：pip3 install deep-translator")
        return lines

    translator = GoogleTranslator(source="auto", target="zh-CN")
    translated = []

    for i in range(0, len(lines), batch_size):
        batch = lines[i:i + batch_size]
        # 用特殊分隔符合并，翻译后拆分
        separator = " \n⟪SEP⟫\n "
        merged = separator.join(batch)
        try:
            result = translator.translate(merged)
            parts = result.split("⟪SEP⟫")
            # 清理多余空白
            parts = [p.strip() for p in parts]
            # 如果拆分数量不匹配，回退为逐条翻译
            if len(parts) != len(batch):
                for line in batch:
                    try:
                        translated.append(translator.translate(line).strip())
                    except Exception:
                        translated.append(line)
            else:
                translated.extend(parts)
        except Exception as e:
            warn(f"批量翻译失败，回退逐条翻译：{e}")
            for line in batch:
                try:
                    translated.append(translator.translate(line).strip())
                except Exception:
                    translated.append(line)

    return translated


def info(msg):
    print(f"{GREEN}[✓]{RESET} {msg}")

def warn(msg):
    print(f"{YELLOW}[!]{RESET} {msg}")

def error(msg):
    print(f"{RED}[✗]{RESET} {msg}")

def step(msg):
    print(f"{CYAN}[→]{RESET} {msg}")


# ─── API Key 管理（CLI > 环境变量 > 配置文件）────────────────────────────────────
def get_api_key(cli_key=None):
    if cli_key:
        return cli_key
    env_key = os.environ.get(ENV_KEY)
    if env_key:
        return env_key
    if CONFIG_FILE.exists():
        try:
            data = json.loads(CONFIG_FILE.read_text())
            key = data.get("api_key")
            if key:
                return key
        except (json.JSONDecodeError, OSError):
            pass
    return None


def print_key_guide():
    error("未检测到 RedFox API Key，请任选一种方式配置：")
    print(f"  1. 环境变量（推荐）：export {ENV_KEY}=ak_你的密钥")
    print(f"  2. 命令行参数：--api-key ak_你的密钥")
    print(f"  3. 配置文件：echo '{{\"api_key\":\"ak_你的密钥\"}}' > ~/.qoder/apis/redfox.json")
    print(f"  注册地址：https://redfox.hk/settings/api-keys?source=github")


# ─── 工具函数 ─────────────────────────────────────────────────────────────────────
def fmt_ts(seconds):
    """秒 → MM:SS 或 HH:MM:SS"""
    try:
        seconds = int(float(seconds))
    except (ValueError, TypeError):
        return "00:00"
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def video_url_of(video_id):
    return f"https://www.youtube.com/watch?v={video_id}" if video_id else ""


# ─── 文字稿接口（带递增延迟重试 + 限频处理）──────────────────────────────────────
def fetch_transcript(session, api_key, video_url, language, include_timestamp, send_metadata):
    """
    调用 transcript 接口。
    成功返回 data 字典（含 video_id / language / transcript [/ metadata]），失败返回 None。
    重试覆盖三类失败：网络异常、状态码错误、业务数据为空；鉴权失败不重试。
    """
    payload = {
        "videoUrl": video_url,
        "format": "json",
        "includeTimestamp": include_timestamp,
        "sendMetadata": send_metadata,
        "language": language,
        "source": SOURCE,
    }
    headers = {"REDFOX_API_KEY": api_key, "Content-Type": "application/json"}

    for attempt in range(MAX_RETRIES):
        try:
            resp = session.post(API_URL, json=payload, headers=headers, timeout=30)
            result = resp.json()

            # 网关包装层：{code, data, msg}
            if isinstance(result, dict) and "code" in result:
                code = result.get("code")
                if code in (200, 2000):
                    data = result.get("data") if isinstance(result.get("data"), dict) else result
                elif code in (3106, 3107):
                    error("API Key 无效或已过期，请检查配置（3106/3107）")
                    return None
                elif code == 3108:
                    warn("触发接口限频，5 秒后重试…")
                    time.sleep(5)
                    continue
                else:
                    warn(f"接口返回错误：code={code} msg={result.get('msg', '')}")
                    data = None
            else:
                # 未包装，直接是文字稿结构
                data = result

            if isinstance(data, dict):
                transcript = data.get("transcript")
                # text 格式兜底：纯文本字符串也算成功
                if transcript or data.get("text"):
                    return data
            warn("返回数据为空或不含文字稿（视频可能无字幕）")

        except requests.exceptions.RequestException as e:
            warn(f"网络请求异常：{e}")
        except (json.JSONDecodeError, ValueError):
            warn("响应解析失败（非 JSON）")

        if attempt < MAX_RETRIES - 1:
            delay = 0.5 * (attempt + 1)  # 0.5s → 1.0s 递增
            time.sleep(delay)

    return None


# ─── 输出格式化 ─────────────────────────────────────────────────────────────────────
def render_lines(data, with_timestamp=True):
    """把 transcript 渲染成行列表"""
    lines = []
    transcript = data.get("transcript") or []
    for seg in transcript:
        text = (seg.get("text") or "").strip()
        if not text:
            continue
        if with_timestamp:
            lines.append(f"[{fmt_ts(seg.get('start', 0))}] {text}")
        else:
            lines.append(text)
    if not lines and data.get("text"):
        lines = [t.strip() for t in str(data["text"]).splitlines() if t.strip()]
    return lines


def total_duration_of(data):
    transcript = data.get("transcript") or []
    if not transcript:
        return 0
    last = transcript[-1]
    try:
        return float(last.get("start", 0)) + float(last.get("duration", 0))
    except (ValueError, TypeError):
        return 0


def build_markdown(data, lines, with_timestamp, video_url_input):
    video_id = data.get("video_id", "")
    language = data.get("language", "")
    metadata = data.get("metadata") or {}
    title = metadata.get("title") or data.get("title") or ""
    author = metadata.get("author") or metadata.get("channel") or data.get("author") or ""

    md = ["# YouTube 视频文案", ""]
    md.append("| 项目 | 内容 |")
    md.append("|------|------|")
    if title:
        md.append(f"| 标题 | {title} |")
    if author:
        md.append(f"| 频道 | {author} |")
    md.append(f"| 视频ID | {video_id} |")
    md.append(f"| 视频链接 | {video_url_of(video_id) or video_url_input} |")
    md.append(f"| 语言 | {language} |")
    md.append(f"| 片段数 | {len(data.get('transcript') or [])} |")
    md.append(f"| 总时长 | {fmt_ts(total_duration_of(data))} |")
    md.append(f"| 时间戳 | {'含' if with_timestamp else '不含'} |")
    md.append(f"| 提取时间 | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} |")
    md.append("")
    md.append("---")
    md.append("")
    md.extend(lines)
    md.append("")
    return "\n".join(md)


# ─── Excel 导出 ──────────────────────────────────────────────────────────────────
def export_excel(data, lines, fpath):
    """导出 xlsx：标题 / 时长 / 视频链接 / 文案内容（与当前时间戳模式一致）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        return False

    video_id = data.get("video_id", "")
    metadata = data.get("metadata") or {}
    title = metadata.get("title") or data.get("title") or video_id

    wb = Workbook()
    ws = wb.active
    ws.title = "视频文案"
    ws.append(["标题", "时长", "视频链接", "文案内容"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append([title, fmt_ts(total_duration_of(data)), video_url_of(video_id), "\n".join(lines)])

    for col, width in {"A": 40, "B": 10, "C": 52, "D": 100}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(fpath)
    return True


# ─── 主流程 ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="YouTube 提文案 — 输入 YouTube 链接/ID 提取视频文字稿"
    )
    parser.add_argument("video_url", help="YouTube 视频 URL（完整/短链）或直接给视频 ID")
    parser.add_argument("--language", default=DEFAULT_LANGUAGE,
                        help=f"字幕语言优先级，逗号分隔（默认 {DEFAULT_LANGUAGE}：中文字幕轨优先，无中文回退英文，asr=自动生成字幕兜底；接口仅选轨不翻译）")
    parser.add_argument("--timestamp", action="store_true",
                        help="输出带 [MM:SS] 时间戳的文案（默认不带时间戳）")
    parser.add_argument("--no-metadata", action="store_true",
                        help="不获取视频元数据（默认获取标题/频道）")
    parser.add_argument("--excel", action="store_true",
                        help="同时导出 Excel（.xlsx，需 openpyxl）")
    parser.add_argument("--json", action="store_true",
                        help="终端输出原始 JSON 响应")
    parser.add_argument("--no-save", action="store_true",
                        help="不保存 Markdown 文件")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR),
                        help=f"输出目录（默认 {DEFAULT_OUTPUT_DIR}）")
    parser.add_argument("--no-translate", dest="translate", action="store_false",
                        help="禁用自动翻译（默认非中文字幕自动翻译为中文）")
    parser.set_defaults(translate=True)
    parser.add_argument("--api-key", default=None, help="指定 RedFox API Key")
    args = parser.parse_args()

    if not HAS_REQUESTS:
        error("缺少 requests 库，请先执行：pip3 install requests")
        sys.exit(1)

    api_key = get_api_key(args.api_key)
    if not api_key:
        print_key_guide()
        sys.exit(1)

    with_timestamp = args.timestamp
    step(f"正在提取文字稿：{args.video_url}（语言优先级：{args.language}）")

    session = requests.Session()
    data = fetch_transcript(
        session, api_key, args.video_url,
        language=args.language,
        include_timestamp=True,  # 始终请求时间戳用于计算总时长，展示层控制是否渲染
        send_metadata=not args.no_metadata,
    )
    if data is None:
        error("提取失败：多次重试后仍无法获取文字稿")
        error("可能原因：视频无字幕 / 视频不存在或受限 / API 额度不足")
        sys.exit(1)

    if args.json:
        print(json.dumps(data, ensure_ascii=False, indent=2))

    # ─── 概要 ───
    video_id = data.get("video_id", "")
    language = data.get("language", "")
    transcript = data.get("transcript") or []
    metadata = data.get("metadata") or {}
    title = metadata.get("title") or data.get("title") or ""
    author = metadata.get("author") or metadata.get("channel") or data.get("author") or ""

    info(f"提取成功：视频 {video_id}")
    if title:
        print(f"  标题：{title}")
    if author:
        print(f"  频道：{author}")
    print(f"  语言：{language} ｜ 片段数：{len(transcript)} ｜ 总时长：{fmt_ts(total_duration_of(data))}")
    print(f"  链接：{video_url_of(video_id) or args.video_url}")

    # ─── 自动翻译（非中文字幕 → 中文）───
    need_translate = args.translate and not is_chinese(language)
    if need_translate:
        step(f"检测到字幕语言为 {language}，正在自动翻译为中文…")

    # ─── 文案全文 ───
    lines = render_lines(data, with_timestamp)
    if not lines:
        warn("文字稿为空，无文案可展示")
    else:
        if need_translate:
            lines = translate_lines(lines)
            info(f"翻译完成（{len(lines)} 段）")
        print(f"\n{DIM}{'─' * 60}{RESET}")
        print(f"{BOLD}文案全文（{len(lines)} 段）：{RESET}\n")
        for line in lines:
            print(line)
        print(f"{DIM}{'─' * 60}{RESET}\n")

    # ─── 保存文件 ───
    if not args.no_save:
        out_dir = Path(args.output_dir).expanduser()
        out_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"{video_id or 'transcript'}_{stamp}.md"
        fpath = out_dir / fname
        fpath.write_text(build_markdown(data, lines, with_timestamp, args.video_url), encoding="utf-8")
        info(f"文案已保存：{fpath}")

    # ─── Excel 导出 ───
    if args.excel:
        out_dir = Path(args.output_dir).expanduser()
        out_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xpath = out_dir / f"{video_id or 'transcript'}_{stamp}.xlsx"
        if export_excel(data, lines, xpath):
            info(f"Excel 已保存：{xpath}")
        else:
            warn("未安装 openpyxl，无法导出 Excel，请执行：pip3 install openpyxl")


if __name__ == "__main__":
    main()
